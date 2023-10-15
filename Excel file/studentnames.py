
import openpyxl
 

path ="C:\\Users\\\Acer\\Documents\\try.xlsx"
 

wb_obj = openpyxl.load_workbook(path)
 
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
 


for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 5)
    print(cell_obj.value)
