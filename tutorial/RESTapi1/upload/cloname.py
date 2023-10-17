# importing openpyxl module
import openpyxl

# Give the location of the file
path = "c:\\Users\\Admin\\Desktop\\she1.xlsx"
 

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
max_col = sheet_obj.max_column

# Loop will print all columns name
for i in range(1, max_col + 1):
	cell_obj = sheet_obj.cell(row = 1, column = i)
	print(cell_obj.value)
