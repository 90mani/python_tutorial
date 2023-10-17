#import openpyxl module

import openpyxl

#call a workbook() function of openpyxl
#to create a new blank workbook object
wb = openpyxl.Workbook()

#Get workbook active sheet
#from the active attribute
sheet = wb.active

#cell objects also have row, column
#and coordinate Attributes that provide
#location information for the cell.

#Note: the first row or column integer
# is 1,not 0 .cell object is created by
# using sheet object's cell()
c1 = sheet.cell(row = 1, column = 1) 

#writing values to cells
c1.value = "Bhuvana"

c2 =sheet.cell(row = 1, column = 2)
c2.value = "Anand"
#once have the Worksheet object, one can
# access a cell object by its name also.
#A2 means column = 1 & row = 2.
c3 =sheet['A2']
c3.value = "Charu"
#B2 means column = 2 & row = 2.
c4 =sheet['B2']
c4.value = "Anand"

#Anytime you modify the workbook object
#or its sheets and cells, the spreadsheet
#file will not be saved until you call 
#the save() workbook method.
wb.save("c:\\Users\\Admin\\Desktop\\she1.xlsx")